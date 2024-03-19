import urllib.request, urllib.parse, urllib.error
from lxml import etree
import re
from openpyxl import Workbook, load_workbook
from sys import exit
import datetime

now=datetime.datetime.now()

ns={'dc':'http://purl.org/dc/elements/1.1/', 'oai_dc':'http://www.openarchives.org/OAI/2.0/oai_dc/'} #definieren der Namespaces für das XML-Parsing

while True:

    ID = input('ID eingeben oder Enter zum beenden: ')

    if len(ID)<1:
        exit()

    xlsx=load_workbook('uscholar checks author upload.xlsx') #Tabelle laden
    sheet=xlsx.active

    newrow=int(sheet.max_row)+1 #Bestimmung der nächsten Zeile

    IDexists = False #Boole'sche Variable, die aussagt, ob die ID bereits in der Tabelle vorhanden ist, damit keine Dubletten erzeugt werden

    for cellObj in sheet['A1':str('A'+str(sheet.max_row))]: #Prüfung auf existieren der ID, verschachtelter Loop
        for cell in cellObj:
            if str(ID) in str(cell.value):
                IDexists = True
            else:
                break
        else:
            break

    if IDexists != False: #Nur Daten lesen und schreiben, wenn ID noch nicht in der Tabelle existiert
        print('ID existiert bereits, Eingabefehler?')
        continue
    else:
        url='https://services.phaidra.univie.ac.at/api/object/o:'+str(ID)+'/index/dc'

    try: #Wenn ID invalide ist, wird Fehlermeldung erzeugt
        data = urllib.request.urlopen(url).read()

    except:
        print('Keine valide ID, Eingabefehler?')
        continue
    
    my_parser=etree.XMLParser(recover=True) #Steuerzeichen filtern
    tree = etree.fromstring(data, parser=my_parser)
    #tree=etree.tostring(xml)

    identifiers=tree.findall('dc:identifier',ns)

    for identifier in identifiers: #PhaidraID finden
        if 'https://phaidra.univie.ac.at/o:' not in identifier.text:
            continue
        else:
            PhaidraID=re.findall('https://phaidra.univie.ac.at/o:(.+)',identifier.text)[0]

    issn='' #Damit ISSN leer wenn nicht vorhanden

    sources=tree.findall('dc:source',ns) #Zeitschrift finden, leer wenn nicht vorhanden
    if len(sources)<1:
        ErschienenIn=''
    else:
        for source in sources:
            if 'issn:' in source.text:
                issn=re.findall('issn:(.+)',source.text)[0] #ISSN für Sherpa-Abfrage
                continue
            else:
                #issn='' eigentlich falsch an dieser Stelle?!
                ErschienenIn=str(source.text)

    EmbargoEnde='' #Damit Embargo-Enddatum leer wenn nicht vorhanden

    dates=tree.findall('dc:date',ns) #Datum finden
    if not dates:
        Erscheinungsdatum='' #Damit Datum leer wenn nicht vorhanden
    else:
        for date in dates:
            if 'info:eu-repo/date/embargoEnd/' in date.text:
                EmbargoEnde=re.findall('info:eu-repo/date/embargoEnd/(.+)',date.text)[0] #Embargo-Enddatum holen
                continue
            else:
                Erscheinungsdatum=str(date.text)

    dc_metadata={'N':'dc:creator', 'O':'dc:title', 'S':'dc:publisher'} #Dictionary für Elemente und Spalten

    xlsx_metadata={'R':Erscheinungsdatum,'U':EmbargoEnde,'A':PhaidraID,'Q':ErschienenIn,'B':'ACCEPTANCE','C':'PENDING','G':str(now)[:10],'AG':'=HYPERLINK(CONCATENATE("https://uscholar.univie.ac.at/o:",A'+str(newrow)+'))',
                   'AH':'=HYPERLINK(CONCATENATE("https://redmine.phaidra.org/redmine/issues/",L'+str(newrow)+'))',
                   'AI':'=CONCATENATE("o",A'+str(newrow)+'," - ",E'+str(newrow)+'," - ",O'+str(newrow)+'," - ",C'+str(newrow)+')',
                   'AJ':'=CONCATENATE("10.25365/phaidra.",M'+str(newrow)+'," > https://phaidra.univie.ac.at/o:",A'+str(newrow)+'," OK")'} #Dictionary für Tabelle, mit fixen Werten und Formeln vorausgefüllt

    accessrights={'metadata only access':'Gesperrter Zugang','embargoed access':'Embargo','restricted access':'Beschränkter Zugang','open access':'Frei zugänglich'} #Dictionary für Rechte

    rights=tree.findall('dc:rights',ns)

    for right in rights: #Zugriffsrechte finden
        if any(c in right.text for c in('All rights reserved','CC','http://creativecommons.org/licenses')): #diese Angaben ausschließen
            continue
        else:
            accesscond=str(right.text)

    if accesscond in accessrights: #Rechtsbegriff für Tabelle einfügen
        xlsx_metadata['T']=accessrights[accesscond]
    else:
        xlsx_metadata['T']=''

    if len(issn)<1: #Sherpa-Link erzeugen, mit ISSN wenn vorhanden
        xlsx_metadata['AK']='=HYPERLINK(CONCATENATE("https://v2.sherpa.ac.uk/cgi/search/publication/basic?publication_title-auto=",Q'+str(newrow)+'))' #warum nicht mit Variable wie bei issn?
    else:
        xlsx_metadata['AK']='=HYPERLINK("https://v2.sherpa.ac.uk/cgi/search/publication/basic?publication_title-auto='+str(issn)+'")'

    for k,v in dc_metadata.items(): #Mapping vom Inhalt der dc-Elemente auf die Tabellenspalten ins Tabellendictionary
        try:
            xlsxk=k
            k=str(tree.find(v,ns).text)
            #print(k)
            xlsx_metadata[xlsxk]=k
        except:
            continue

    for k,v in xlsx_metadata.items(): #Befüllung der Tabelle mittels Dictionary
        try:
            pos=str(k)+str(newrow)
            #print(pos)
            sheet[pos]=v
            #print(sheet[pos].value)
        except:
            continue

    xlsx.save('uscholar checks author upload.xlsx')

    print('o:'+str(ID)+' erfolgreich in Zeile '+str(newrow)+' geschrieben!') #Erfolgsbestätigung
